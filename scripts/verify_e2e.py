"""Verify the spreadsheet the PACKAGED app produced, figure by figure.

This runs inside the build job (not the test job) so that a missing module
or a broken bundle fails the build loudly instead of reaching a user.
"""
import glob
import sys
from decimal import Decimal

import openpyxl

workdir = sys.argv[1]
outs = glob.glob(f"{workdir}/output/*.xlsx")
assert len(outs) == 1, f"expected exactly one output workbook, got {outs}"

wb = openpyxl.load_workbook(outs[0])
ws = wb["Data"]
rows = list(ws.iter_rows(min_row=2, values_only=True))
header = [c.value for c in ws[1]]
i_amt, i_status = header.index("Amount"), header.index("Status")
i_date, i_type = header.index("Date"), header.index("Type")

amounts = [Decimal(str(r[i_amt])) for r in rows]
expected = [Decimal("500.25"), Decimal("200.25"),
            Decimal("-150.25"), Decimal("-50.00")]
assert amounts == expected, f"amounts wrong: {amounts}"
assert all(r[i_status] == "OK" for r in rows), "rows flagged on a clean file"
assert all(r[i_date] is not None for r in rows), "dates missing"
types = [r[i_type] for r in rows]
assert types == ["Credit", "Credit", "Debit", "Debit"], types

ri = wb["Run Info"]
info = {r[0].value: r[1].value for r in ri.iter_rows()}
assert info["Reconciliation"] == "PASSED", info["Reconciliation"]
assert "1,000.00" in info["Reconciliation detail"]
assert "1,500.25" in info["Reconciliation detail"]

# the safety net must be real: a clean run proves nothing unless a broken
# run fails (the deliberate-breakage check lives in the workflow)
print("E2E VERIFY OK — 4 rows, every figure exact, reconciliation PASSED")

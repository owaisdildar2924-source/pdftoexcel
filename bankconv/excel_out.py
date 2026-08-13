"""Excel output: one workbook per PDF, Data + Run Info sheets.

Matches the schema the client already uses. Real Excel dates and numbers,
never text. NEEDS_REVIEW rows highlighted, with the specific reason in the
Review Note column. A Raw Extracted Text column appears only when at least
one row needs it.
"""
from __future__ import annotations

import datetime as _dt
from decimal import Decimal
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
REVIEW_FILL = PatternFill("solid", fgColor="FFE699")
DATE_FMT = "yyyy-mm-dd"
MONEY_FMT = "#,##0.00"

BANK_COLS = ["Source File", "Institution", "Account", "Date", "Description",
             "Type", "Amount", "Balance", "Status", "Review Note"]
SETTLE_COLS = ["Section", "Job #", "Customer", "Type", "Description",
               "Amount", "Status", "Review Note"]


def _style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"


def _autosize(ws, widths: dict[int, int]) -> None:
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


def _num(cell, value: Optional[Decimal]) -> None:
    if value is not None:
        cell.value = float(value)
        cell.number_format = MONEY_FMT


def write_bank_workbook(path: str, source_name: str, institution: str,
                        account: Optional[str], blocks, run_info: dict) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    rows = [(b, t) for b in blocks for t in b.txns]
    need_raw = any(t.status != "OK" for _, t in rows)
    cols = BANK_COLS + (["Raw Extracted Text"] if need_raw else [])
    ws.append(cols)
    _style_header(ws, len(cols))

    for b, t in rows:
        r = ws.max_row + 1
        amount = t.amount
        if amount is not None and b.is_credit_card:
            amount = -amount     # cash perspective: purchases negative
        typ = ""
        if amount is not None:
            typ = "Credit" if amount >= 0 else "Debit"
        acct = b.name if (b.name and not b.name.lower().startswith(
            ("beginning", "balance"))) else (account or "")
        ws.cell(r, 1, source_name)
        ws.cell(r, 2, institution)
        ws.cell(r, 3, acct)
        dcell = ws.cell(r, 4)
        if t.date:
            dcell.value = t.date
            dcell.number_format = DATE_FMT
        ws.cell(r, 5, t.description[:500])
        ws.cell(r, 6, typ)
        _num(ws.cell(r, 7), amount)
        _num(ws.cell(r, 8), t.balance)
        ws.cell(r, 9, t.status)
        ws.cell(r, 10, t.review_note or None)
        if need_raw:
            ws.cell(r, 11, t.raw_text[:500] if t.status != "OK" else None)
        if t.status != "OK":
            for c in range(1, len(cols) + 1):
                ws.cell(r, c).fill = REVIEW_FILL
    _autosize(ws, {1: 28, 2: 18, 3: 20, 4: 12, 5: 60, 6: 8, 7: 12, 8: 12,
                   9: 14, 10: 46, 11: 50})
    _write_run_info(wb, run_info)
    wb.save(path)


def write_settlement_workbook(path: str, rows, run_info: dict) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    need_raw = any(x.status != "OK" for x in rows)
    cols = SETTLE_COLS + (["Raw Extracted Text"] if need_raw else [])
    ws.append(cols)
    _style_header(ws, len(cols))
    for x in rows:
        r = ws.max_row + 1
        ws.cell(r, 1, x.section)
        ws.cell(r, 2, x.job)
        ws.cell(r, 3, x.customer)
        ws.cell(r, 4, x.rtype)
        ws.cell(r, 5, x.description[:400])
        _num(ws.cell(r, 6), x.amount)
        ws.cell(r, 7, x.status)
        ws.cell(r, 8, x.review_note or None)
        if need_raw:
            ws.cell(r, 9, x.raw[:400] if x.status != "OK" else None)
        if x.status != "OK":
            for c in range(1, len(cols) + 1):
                ws.cell(r, c).fill = REVIEW_FILL
    _autosize(ws, {1: 12, 2: 14, 3: 22, 4: 12, 5: 55, 6: 12, 7: 14, 8: 46, 9: 50})
    _write_run_info(wb, run_info)
    wb.save(path)


RECON_RULE = ("Where the statement prints a running balance, every row must "
              "move that balance by exactly its own amount, within 1 cent. "
              "Where it does not, the printed beginning balance plus all "
              "movements must equal the printed ending balance. Settlement "
              "statements are checked against every printed subtotal and "
              "total instead.")


def _write_run_info(wb: Workbook, info: dict) -> None:
    ws = wb.create_sheet("Run Info")
    pairs = [
        ("Source file", info.get("source", "")),
        ("Format detected", info.get("format", "")),
        ("Detection confidence", info.get("confidence", "")),
        ("Extraction method", info.get("method", "")),
        ("Pages read", info.get("pages", "n/a")),
        ("OCR confidence", info.get("ocr_conf", "n/a")),
        ("Rows extracted", info.get("rows", 0)),
        ("Rows needing review", info.get("review", 0)),
        ("Rows failed", info.get("failed", 0)),
        ("Reconciliation", info.get("recon", "")),
        ("Reconciliation detail", info.get("recon_detail", "")),
        ("Reconciliation rule", RECON_RULE),
        ("Processed at", _dt.datetime.now().replace(microsecond=0)),
        ("Produced by", info.get("producer", "PDF to Excel 2.0.0")),
    ]
    for k, v in pairs:
        ws.append([k, v])
    for row in ws.iter_rows(min_col=1, max_col=1):
        row[0].font = Font(bold=True)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 110
    for row in ws.iter_rows(min_col=2, max_col=2):
        row[0].alignment = Alignment(wrap_text=False, vertical="top")
